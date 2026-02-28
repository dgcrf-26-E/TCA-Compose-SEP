from django.shortcuts import render, redirect, get_object_or_404, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import ListView, CreateView, UpdateView
from django.contrib.auth.models import User
from django.db.models import Max
# from usuarios.models import usuarioL
from usuarios.models import UsuarioP
from usuarios.models import Registro, Acciones, Notificacion,Area, Rubro, Mensaje
from datetime import datetime, timedelta, timezone
from .forms import RegistroConAccionesYPruebasForm, MensajeForm, AccionesForm, RegistroConAccionesFORM, CargarArchivoForm
from django.forms import inlineformset_factory
import openpyxl as opxl
from django.core.paginator import Paginator

import locale
from datetime import date
from datetime import datetime
import re
from django.contrib import messages

from django.db import transaction
from dateutil.parser import parse

import os
import zipfile
from io import BytesIO

from django.http import JsonResponse
from django.urls import reverse_lazy

# from twilio.rest import Client

# import os
# from dotenv import load_dotenv

from django.db.models.functions import Substr, StrIndex, Reverse, Length
from django.db.models import Value, F

# load_dotenv()
# registros_con_multiples_acciones = Acciones.objects.values('idRegistro', 'idAccion').annotate( total_acciones=Count('idAccion') ).filter(total_acciones__gte=2)

@login_required
def dashboard(request):
    if request.method == 'GET':

        filtro = request.GET.get('filtro')
        tiempo = request.GET.get('año')
        respA = request.GET.get('resp')
        
        # print(f"filtro {filtro}")
        # print(f"filtro {tiempo}")

        consultarAreas = Area.objects.all()
        lista_años = [str(año) for año in range(2020, datetime.now().year + 1)]

        userDataI = UsuarioP.objects.filter(user__username=request.user)
        registrosConFechas = []
        nombres_areas = [area.nickname for area in consultarAreas]
        areas_n = nombres_areas if len(nombres_areas) > 1 else None

        registros = Registro.objects.all()

        # print(areas_n)
        formCargar1 = CargarArchivoForm()
        if userDataI[0].tipo == "1":
            if filtro in nombres_areas:
                filtroB = Area.objects.get(nickname = filtro)
                registros = registros.filter(area=filtroB.idArea)

            if tiempo in lista_años:
                print(f"filtro {tiempo}")
                registros = registros.filter(fecha_inicio__year=tiempo)

            if respA in nombres_areas:
                filtroC = Area.objects.get(nickname = respA)
                registros = registros.filter(accionR__area2=filtroC)
            
            # print(f"registros {registros.count()}")
        else:
            registros_area = Registro.objects.filter(area=userDataI[0].OR)
            registros_acciones_area2 = Registro.objects.filter(
                accionR__area2=userDataI[0].OR
            )

            registros = registros_area | registros_acciones_area2
            registros = registros.distinct()

            if filtro in nombres_areas:
                filtroB = Area.objects.get(nickname = filtro)
                registros = registros.filter(area=filtroB.idArea)

            if tiempo in lista_años:
                # print(f"filtro {tiempo}")
                registros = registros.filter(fecha_inicio__year=tiempo)

            if respA in nombres_areas:
                filtroC = Area.objects.get(nickname = respA)
                registros = registros.filter(accionR__area2=filtroC)
            
            # print(f"registros {registros.count()}")

        # filtrar por registros atendidos y no atendidos
        # registros_en_proceso = registros.filter(estado="1")
        # registros_atendidos = registros.filter(estado="2")

        # registros_ordenados = list(registros_en_proceso) + list(registros_atendidos)

        registros_en_proceso = registros.annotate(
            year = Substr('claveAcuerdo', Length('claveAcuerdo') - 4, 4),
            oficina = Substr("claveAcuerdo", 4, 4),
            mes=Substr('claveAcuerdo', Length('claveAcuerdo') - 6, 2),  # Extrae el mes (asumiendo que está en la posición 10-11)
            numero=Substr('claveAcuerdo', 1, 3)  # Extrae el número (001, 002, etc.)
        ).order_by('oficina', 'year', 'mes', 'numero')

        # registros_ordenados = registros.annotate(
        #     segundo_slash=StrIndex('claveAcuerdo', Value('/', 2)),
        #     mes=Substr('claveAcuerdo', F('segundo_slash') + 1, 2),  # Extrae el mes (asumiendo que está en la posición 10-11)
        #     numero=Substr('claveAcuerdo', 1, 3)  # Extrae el número (001, 002, etc.)
        # ).order_by('mes', 'numero')

        # registros_en_proceso = registros.annotate(
        #     # segundo_slash=StrIndex('/20', "claveAcuerdo"),  # Busca la posición del segundo '/'
        #     mes=Substr('claveAcuerdo', Length("claveAcuerdo") - 7 , 2),  # Extrae el mes basándose en el segundo '/'
        #     numero=Substr('claveAcuerdo', 1, 3)  # Extrae el número al principio
        # ).order_by('mes', 'numero')

        registros_ordenados = list(registros_en_proceso)

        for registro in registros_ordenados:
            fecha_inicio_str = registro.fecha_inicio.strftime('%d-%m-%Y')
            fecha_inicio = fecha_inicio_str.split('-')

            fecha_termino_str = registro.fecha_termino.strftime('%d-%m-%Y')
            fecha_termino = fecha_termino_str.split('-')

            fecha_inicio_dt = datetime.strptime(fecha_inicio_str, '%d-%m-%Y')
            fecha_termino_dt = datetime.strptime(fecha_termino_str, '%d-%m-%Y')
            diferencia = datetime.now() - fecha_termino_dt

            fecha_finalizacion = registro.fecha_finalizacion

            areas = registro.area.all()
            areas_str = ', '.join(area.nickname for area in areas)
            areas_name = ', '.join(area.name for area in areas)

            dias = diferencia.days
            porcentaje = registro.porcentaje_avance
            clave_acuerdo_partes = registro.claveAcuerdo.split('/')

            registrosConFechas.append({
                'registro': registro,
                'fecha_inicio': fecha_inicio,
                'fecha_termino': fecha_termino,
                'diferencia': dias,
                'areas_str': areas_str,
                'areas_name': areas_name,
                'fecha_finalizacion': fecha_finalizacion,
                'porcentaje': porcentaje,
                'clave_acuerdo_partes': clave_acuerdo_partes,
            })

        # now = datetime.now()
        # nuevos_registros = registros.filter(fecha_creacion__gte=now - timedelta(days=365))


        # for registro in nuevos_registros:
        #     Notificacion.objects.get_or_create(user=request.user, registro=registro)

        notificaciones = Notificacion.objects.filter(user=request.user, leido=False)

        paginator = Paginator(registrosConFechas, 100)  # Show 200 contacts per page.
        page_number = request.GET.get("page")
        page_obj = paginator.get_page(page_number)

        context = {
            'registrosConFechas': page_obj,
            'dataU': userDataI,
            'notificaciones': notificaciones,
            'formCargar1': formCargar1,
            'areas_n': areas_n,
            'lista_años': lista_años,
            'filtroOR': filtro,
            'filtroAño': tiempo,
            'filtroResp': respA,
        }

        user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
        if 'mobile' in user_agent:
            # template = 'home_mobile.html'
            return render(request, "dashboard/dashboard_m.html", context)
        else:
            return render(request, "dashboard/dashboard.html", context)
            # template = 'home_pc.html'
        return render(request, template)

        # return render(request, "dashboard/dashboard.html", context)

@login_required
def marcar_notificacion_leida(request, notificacion_id):
    notificacion = get_object_or_404(Notificacion, id=notificacion_id, user=request.user)
    notificacion.leido = True
    notificacion.fecha_leido = datetime.now()
    notificacion.save()

    registro_id = 1
    if notificacion.registro_id != 0 and notificacion.registro_id != None :
        registro_id = notificacion.registro_id

    return redirect('detalles', registro_id=registro_id)



@login_required
def detalles(request, registro_id):
    registro = Registro.objects.get(pk=registro_id)
    mensajes = registro.mensajes.all().order_by('fecha_envio')
    userDataI = UsuarioP.objects.filter(user__username=request.user).first()

    for mensaje in mensajes:
        if mensaje.archivo:
            mensaje.archivo_nombre = mensaje.archivo.name.split('/')[-1]

    
    # if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
    #     form = MensajeForm(request.POST, request.FILES)
    #     if form.is_valid():
    #         nuevo_mensaje = form.save(commit=False)
    #         nuevo_mensaje.registro = registro
    #         nuevo_mensaje.usuario = request.user
    #         print(form.save())
    #         # form.save()
    #         return JsonResponse({'message': 'Archivo subido con éxito'}, status=200)
    #     else:
    #         return JsonResponse({'error': form.errors}, status=400)
        
    if request.method == 'POST':
        mensaje_form = MensajeForm(request.POST, request.FILES)
        if mensaje_form.is_valid():

            nuevo_mensaje = mensaje_form.save(commit=False)
            nuevo_mensaje.registro = registro
            nuevo_mensaje.usuario = request.user
            nuevo_mensaje.save()

            mensaje = f"Tienes una notificacion del acuerdo: {registro.claveAcuerdo}."
            
            generarNotificacion(registro.idRegistro, mensaje, userDataI.idUser)

            return redirect('detalles', registro_id=registro_id)
    else:
        mensaje_form = MensajeForm()

    context = {
        'registro': registro,
        'mensajes': mensajes,
        'mensaje_form': mensaje_form,
        'dataU': userDataI,
    }

    return render(request, 'dashboard/detalles.html', context)


@login_required
def descargar_archivos_acuerdo(request, registro_id):
    registro = Registro.objects.get(pk=registro_id)
    claveA = registro.claveAcuerdo
    claveA = claveA.replace('/','-')
    mensajes = registro.mensajes.all().order_by('fecha_envio')
    
    zip_buffer = BytesIO()

    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        for mensaje in mensajes:
            if mensaje.archivo:
                archivo_path = mensaje.archivo.path
                # Agregar el archivo al ZIP, usando el nombre base del archivo
                zip_file.write(archivo_path, os.path.basename(archivo_path))
                
    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer, content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename={claveA}.zip'
    return response

class AcuerdoCreateNew(CreateView):
    model = Registro
    form_class = RegistroConAccionesYPruebasForm
    template_name = 'dashboard/crear_registroN.html'
    success_url = reverse_lazy('dashboard')

@login_required
def crear_registro(request):
    if request.method == 'POST':
        userDataI = UsuarioP.objects.filter(user__username=request.user).first()
        registro_form = RegistroConAccionesYPruebasForm(request.POST)
        if registro_form.is_valid():
            registro = registro_form.save()

            areas2 = registro_form.cleaned_data['accion1_area2']
            area = registro_form.cleaned_data['area']
            claveA = registro_form.cleaned_data['claveAcuerdo']

            accion = Acciones.objects.create(
                descripcion=registro_form.cleaned_data['accion1_descripcion'],
                antecedente=registro_form.cleaned_data['accion1_antecedente']
            )
            
            accion.area2.set(areas2)
            accion.save()
            registro.accionR.add(accion)

            mensaje = f"Se ha creado un nuevo acuerdo: {claveA} en la {area[0].name}."
            generarNotificacion(registro.idRegistro, mensaje, userDataI.idUser)

            return redirect('dashboard')
    else:
        registro_form = RegistroConAccionesYPruebasForm()

    return render(request, 'dashboard/crear_registro.html', {
        'registro_form': registro_form,
    })


@login_required
def editar_registro(request, id):
    registro = get_object_or_404(Registro, idRegistro=id)
    accion = registro.accionR.first()
    userDataI = UsuarioP.objects.filter(user__username=request.user)

    if request.method == 'POST':
        registro_form = RegistroConAccionesFORM(request.POST, instance=registro)
        accion_form = AccionesForm(request.POST, instance=accion)

        if registro_form.is_valid() and accion_form.is_valid():
            # print(accion_form)
            if registro.estado == 1:
                registro.fecha_finalizacion = "1970-01-01"
            else:
                registro.fecha_finalizacion = datetime.now().strftime("%Y-%m-%d")
            registro = registro_form.save()
            accion = accion_form.save()
            accion.registro = registro
            accion.save()

            messages.success(request, "Guardado correctamente.")
            return render(request, 'dashboard/editar_registro.html', {
                'registro_form': registro_form,
                'accion_form': accion_form,
                'dataU': userDataI,

            })

        else:
            for field, errors in registro_form.errors.items():
                for error in errors:
                    messages.error(request, f"Error en {field}: {error}")
            for field, errors in accion_form.errors.items():
                for error in errors:
                    messages.error(request, f"Error en {field}: {error}")
    else:
        registro_form = RegistroConAccionesFORM(instance=registro)
        accion_form = AccionesForm(instance=accion)

    return render(request, 'dashboard/editar_registro.html', {
        'registro_form': registro_form,
        'accion_form': accion_form,
        'dataU': userDataI,
    })


@login_required
def eliminar_registro(request, idRegistro):
    registro = get_object_or_404(Registro, idRegistro=idRegistro)
    accion = registro.accionR.first()
    userDataI = UsuarioP.objects.filter(user__username=request.user)

    if userDataI[0].tipo == "1" and request.method == 'POST':
        registro.delete()
        accion.delete()
        return redirect('dashboard')

    return render(request, 'dashboard/eliminar_registro.html', {'registro': registro})


@login_required
def eliminar_mensaje(request, idMensaje):
    mensajeI = get_object_or_404(Mensaje, id=idMensaje)
    userDataI = UsuarioP.objects.filter(user__username=request.user).first()

    if userDataI.tipo == "1" and request.method == 'POST':
        mensajeI.delete()
        return redirect('detalles', mensajeI.registro.idRegistro)
    else:
        return redirect('detalles', mensajeI.registro.idRegistro)

    # return render(request, 'dashboard/eliminar_registro.html', {'registro': registro})


locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

def convert_spanish_date(date_str):
    try:
        date_str = date_str.lower()
        return datetime.strptime(date_str, '%d de %B de %Y').date()
    except ValueError:
        print(f"Error al convertir fecha: {date_str}")
        return None



@login_required
@csrf_exempt
def cargaMasiva(request):
    if request.method == "POST":
        userDataI = UsuarioP.objects.filter(user__username=request.user).first()
        form = CargarArchivoForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["archivo"]
            nombreA = str(excel_file.name)
            extensionA = (nombreA.split(".")[-1]).lower()

            # print(f"Archivo recibido: {nombreA}")
            errores = 0
            errArea = 0
            if extensionA in ["xlsx", "xlsm", "xlsb", "xltx", "xltm", "xls"]:
                dataWB = opxl.load_workbook(excel_file, data_only=True)
                data = dataWB.worksheets[0]

                claveAcuerdoL = []
                fecha_inicioL = []
                fecha_terminoL = []
                rubroL = []
                areaL = []

                errores = ""
                i = 1
                while not(i == 0):
                    if(data.cell(i+2,2).value == None):
                        i = 0
                    else:
                        
                        claveAcuerdo=data.cell(i+2,3).value
                        # print(claveAcuerdo)
                        fecha_inicio= data.cell(i+2,5).value
                        fecha_termino=data.cell(i+2,4).value
                        estadoA = data.cell(i+2,12).value
                        porcentajeA = data.cell(i+2,13).value
                        try:
                            rubroAux = (data.cell(i+2,8).value)
                            # print(f"rubro {rubroAux}")
                            rubro = Rubro.objects.get(tipo = rubroAux.strip())
                            # rubroL.append(Rubro.objects.get(tipo=rubroAux))
                        except Rubro.DoesNotExist:
                            errores += f"No existe el Rubro {rubroAux}\n"

                        try:
                            areaAux = (data.cell(i+2,6).value)
                            areaL = Area.objects.get(nickname=areaAux)
                        except Area.DoesNotExist:
                            errores += f"Área '{areaAux}' no encontrada"

                        descripcionAux = data.cell(i+2,9).value

                        areas_responsables = data.cell(i+2,10).value
                        # print(areas_responsables)
                        areas_responsables_list = areas_responsables.split('-')
                        areas_objs = []
                        for area in areas_responsables_list:
                            # print(f"responsables {area}")
                            if area.strip() in ["OR", "or", "Or", "OR "]:
                                areas_objs.append(areaL)
                            else:
                                try:
                                    areaAux1 = Area.objects.get(nickname=area.strip())
                                    areas_objs.append(areaAux1)
                                except Area.DoesNotExist:
                                    errores += f"Área '{area}' no encontrada"
                        

                        registro, created = Registro.objects.update_or_create(
                                claveAcuerdo=claveAcuerdo,
                                defaults={
                                    "fecha_inicio": fecha_inicio,
                                    "fecha_termino": fecha_termino,
                                    "estado": str(estadoA),
                                    "porcentaje_avance": porcentajeA,
                                }
                            )
                        
                        registro.area.add(areaL)
                        registro.rubro.add(rubro)
                        registro.save()

                        try:
                            accionesB = Acciones.objects.get(idRegistro=registro)
                            accionesB.area2.aclear()
                            accionesB.descripcion = descripcionAux
                            for area in areas_objs:
                                accionesB.area2.add(area)
                            accionesB.save()

                        except:
                            errores += f"accion '{registro.claveAcuerdo}' no encontrada"

                            accionNueva = Acciones.objects.create(
                                descripcion = descripcionAux
                            )
                            for area in areas_objs:
                                accionNueva.area2.add(area)

                            accionNueva.idRegistro.add(registro)
                            accionNueva.save()

                        # accionNueva = Acciones.objects.create(
                        #     descripcion = descripcionAux
                        # )
                        # for area in areas_objs:
                        #     accionNueva.area2.add(area)

                        # accionNueva.idRegistro.add(registro)
                        # accionNueva.save()

                        mensaje = f"Tienes una notificacion del acuerdo: {claveAcuerdo}."
                        generarNotificacion(registro.idRegistro, mensaje, userDataI.idUser)

                        i+=1
                        # i=0



                # accion, created = Acciones.objects.update_or_create(
                #                 descripcion=descripcion,
                #             )
                #             accion.idRegistro.add(registro)
                #             accion.area2.set(areas_objs)

                        # paises.append([data.cell(i+3,2).value, data.cell(i+3,3).value])

                        
                #-----------------------------------

                #Eliminar todos los registros de Paises
                # Paises.objects.all().delete()
                #------------------------

                # Se agrega a la DB los paises y se crea un string de ellos
                # paisesStr = "{"    
                # for num, nac in enumerate(paises):
                #     paisesStr += f'\n\t"{nac}",' 
                #     Paises.objects.create(idPais = num + 1, nombre_pais = nac[0], iso3 = nac[1])
                #     # print(nac)
                # paisesStr += "\n}" 
                #----------------------------

                # pasar string de las nacionalidades 
                # print(paisesStr) 
                #------------------------

                # total_rows = data.max_row - 1
                # chunk_size = 20

                # for start_row in range(1, total_rows + 1, chunk_size):
                #     end_row = min(start_row + chunk_size, total_rows + 1)
                    
                #     with transaction.atomic():
                #         for i in range(start_row, end_row):
                #             if data.cell(i + 1, 3).value is None:
                #                 break
                #             fecha_inicio = data.cell(i + 1, 1).value
                #             fecha_inicio = convert_spanish_date(fecha_inicio) if isinstance(fecha_inicio, str) else fecha_inicio

                #             area_celda = data.cell(i + 1, 2).value
                #             clave_acuerdo = data.cell(i + 1, 3).value
                #             rubro = data.cell(i + 1, 4).value
                #             descripcion = data.cell(i + 1, 5).value
                #             areas_responsables = data.cell(i + 1, 6).value
                #             fecha_termino = data.cell(i + 1, 7).value
                #             estado = data.cell(i + 1, 8).value
                #             if estado in ["ATENDIDA","Atendida","Atendido", "atendido", "ATENDIDO", "Completado", "COMPLETADO", "Terminado", "TERMINADO", "Cumplido", "CUMPLIDO", "cumplido"]:
                #                 estado = 2
                #             else:
                #                 estado = 1
                #             fecha_finalizacion = data.cell(i + 1, 9).value
                #             if fecha_finalizacion is None:
                #                 fecha_finalizacion = datetime(1970, 1, 1).date()
                #             if not isinstance(fecha_termino, date):
                #                 fecha_termino = fecha_inicio

                #             try:
                #                 rubro_obj = Rubro.objects.get(tipo=rubro)
                #             except Rubro.DoesNotExist:
                #                 errores += 1

                #             try:
                #                 area_obj = Area.objects.get(nickname=area_celda)
                #             except Area.DoesNotExist:
                #                 pass
                #                 print(f"Área '{area_celda}' no encontrada")

                #             areas_responsables_list = areas_responsables.split('-')
                #             areas_objs = []
                #             for area in areas_responsables_list:
                #                 try:
                #                     if area == "OR":
                #                         area_responsable_obj = Area.objects.get(nickname=area_celda)
                #                     else:
                #                         area_responsable_obj = Area.objects.get(nickname=area.strip())
                #                     print(f"Área responsable seteada '{area_responsable_obj}'")
                #                 except Area.DoesNotExist:
                #                     errArea += 1
                #                     print(f"Área responsable '{area_responsable_obj}' no encontrada")
                #                     pass
                #                 areas_objs.append(area_responsable_obj)

                #             ultimo_clave_acuerdo = Registro.objects.filter(area=area_obj).aggregate(Max('claveAcuerdo'))['claveAcuerdo__max']
                #             if ultimo_clave_acuerdo:
                #                 ultimo_indice = int(ultimo_clave_acuerdo.split('/')[0])
                #             else:
                #                 ultimo_indice = 0
                #             nuevo_indice = ultimo_indice + 1
                #             clave_acuerdo = f"{nuevo_indice:02}/{area_celda.split(' ')[1]}/{fecha_inicio.strftime('%m/%Y')}"

                #             registro, created = Registro.objects.update_or_create(
                #                 claveAcuerdo=clave_acuerdo,
                #                 defaults={
                #                     "fecha_inicio": fecha_inicio,
                #                     "fecha_termino": fecha_termino,
                #                     "estado": estado,
                #                     "fecha_finalizacion": fecha_finalizacion,
                #                 }
                #             )
                #             registro.rubro.set([rubro_obj])
                #             registro.area.set([area_obj])

                #             accion, created = Acciones.objects.update_or_create(
                #                 descripcion=descripcion,
                #             )
                #             accion.idRegistro.add(registro)
                #             accion.area2.set(areas_objs)

                #             print(f"Registro {i}: {clave_acuerdo} procesado correctamente")
                #             print('Errores de area', errArea)

            return redirect('dashboard')
        else:
            print("Formulario no válido")
    else:
        form = CargarArchivoForm()
    return render(request, "dashboard/dashboard.html", {"form": form})


def paginarRegistros(request):
    if request.method == 'GET':
        userDataI = UsuarioP.objects.filter(user__username=request.user)
        registrosConFechas = []
        nombres_areas = [area.nickname for area in Area.objects.all()]
        areas_n = nombres_areas if len(nombres_areas) > 1 else None
    

        # print(areas_n)
        formCargar1 = CargarArchivoForm()
        if userDataI[0].tipo == "1":
            registros = Registro.objects.all().order_by('fecha_termino')
        else:
            registros_area = Registro.objects.filter(area=userDataI[0].OR).order_by('fecha_termino')
            registros_acciones_area2 = Registro.objects.filter(
                accionR__area2=userDataI[0].OR
            ).order_by('fecha_termino')

            registros = registros_area | registros_acciones_area2
            registros = registros.distinct().order_by('fecha_termino')

        registros_en_proceso = registros.filter(estado="1")
        registros_atendidos = registros.filter(estado="2")

        registros_ordenados = list(registros_en_proceso) + list(registros_atendidos)

        for registro in registros_ordenados:
            fecha_inicio_str = registro.fecha_inicio.strftime('%d-%m-%Y')
            fecha_inicio = fecha_inicio_str.split('-')

            fecha_termino_str = registro.fecha_termino.strftime('%d-%m-%Y')
            fecha_termino = fecha_termino_str.split('-')

            fecha_inicio_dt = datetime.strptime(fecha_inicio_str, '%d-%m-%Y')
            fecha_termino_dt = datetime.strptime(fecha_termino_str, '%d-%m-%Y')
            diferencia = datetime.now() - fecha_termino_dt

            fecha_finalizacion = registro.fecha_finalizacion

            areas = registro.area.all()
            areas_str = ', '.join(area.nickname for area in areas)
            areas_name = ', '.join(area.name for area in areas)

            dias = diferencia.days
            porcentaje = registro.porcentaje_avance
            clave_acuerdo_partes = registro.claveAcuerdo.split('/')


            registrosConFechas.append({
                'registro': registro,
                'fecha_inicio': fecha_inicio,
                'fecha_termino': fecha_termino,
                'diferencia': dias,
                'areas_str': areas_str,
                'areas_name': areas_name,
                'fecha_finalizacion': fecha_finalizacion,
                'porcentaje': porcentaje,
                'clave_acuerdo_partes': clave_acuerdo_partes
            })

        now = datetime.now()
        nuevos_registros = registros.filter(fecha_creacion__gte=now - timedelta(days=365))


        for registro in nuevos_registros:
            Notificacion.objects.get_or_create(user=request.user, registro=registro)

        notificaciones = Notificacion.objects.filter(user=request.user, leido=False)

        paginator = Paginator(registrosConFechas, 150)  # Show 200 contacts per page.
        page_number = request.GET.get("page")
        page_obj = paginator.get_page(page_number)

        context = {
            'registrosConFechas': page_obj,
            'dataU': userDataI,
            'notificaciones': notificaciones,
            'formCargar1': formCargar1,
            'areas_n': areas_n
        }

        return render(request, "dashboard/list.html", context=context)

def generarNotificacion(idRegistro, mensaje, idUser):
    # mensaje = f"Tienes una notificacion del acuerdo: {registroI.claveAcuerdo}."
    filtrarA = []
    registroI = Registro.objects.filter(idRegistro=idRegistro).first()

    # Obtener todos los usuarios del área asociada
    accionAcuerdo = Acciones.objects.filter(idRegistro=idRegistro).first()
    
    for area in accionAcuerdo.area2.all():
        auxD = str(area.idArea)
        if auxD not in filtrarA:
            filtrarA.append(auxD)

    # print(areaDGCOR)
    areaDGCOR = Area.objects.filter(nickname="DGCOR").first()
    auxDG = str(areaDGCOR.idArea)
    if auxDG not in filtrarA:
        filtrarA.append(auxDG)
    # print(filtrarA)

    for areas_enOR in registroI.area.all():
        print(areas_enOR)
        auxOR = str(areas_enOR.idArea)
        if auxOR not in filtrarA:
            filtrarA.append(auxOR)

    usuarios_del_area = UsuarioP.objects.filter(OR__in=filtrarA)
    # print(usuarios_del_area)

    usuarios_del_area = usuarios_del_area.exclude(idUser=idUser)

    # Crear una notificación para cada usuario del área
    for usuarioC in usuarios_del_area:

        noti = Notificacion.objects.filter(mensaje=mensaje,user=usuarioC.user).last()

        if noti != None:
            noti.leido = False
            noti.save()
        else:
            Notificacion.objects.create(user=usuarioC.user, mensaje=mensaje, leido = False, registro_id=idRegistro)

        # notifi, created = Notificacion.objects.update_or_create(
        #     mensaje=mensaje,
        #     user=usuarioC.user,
        #     defaults={
        #         "mensaje": mensaje,
        #         "leido": False,
        #         "registro_id": idRegistro,
        #     }
        # )
        # print(notifi, created)

